import json
import openpyxl

json_file_path = "/home/serverguy/Downloads/29_aug.json"

# Read data from the JSON file
with open(json_file_path, "r") as json_file:
    data = json.load(json_file)["items"]

# Extracted data list for text file
extracted_data_for_text = []

workbook = openpyxl.Workbook()
sheet = workbook.active

# Write headers (keys) as the first row
header_row = 1
headers = data[0].keys()
for col_index, key in enumerate(headers, start=1):
    sheet.cell(row=header_row, column=col_index, value=key)

# Write the JSON data to Excel rows
for row_index, item in enumerate(data, start=header_row + 1):
    for col_index, value in enumerate(item.values(), start=1):
        if isinstance(value, list):  # Handle list values
            value = ', '.join(map(str, value))  # Convert list items to strings
        elif isinstance(value, dict):  # Handle dictionary values
            value = json.dumps(value)  # Convert dictionary to JSON string
        sheet.cell(row=row_index, column=col_index, value=value)
    
    extracted_data_for_text.append((item.get("name", ""), item.get("last_backup_created", "")))

output_excel_path = "/home/serverguy/today2_output.xlsx"
workbook.save(output_excel_path)
print("Excel file saved successfully.")

# Save extracted data to a text file
text_file_path = "/home/serverguy/abc.txt"
with open(text_file_path, "w") as f:
    for name, last_backup_created in extracted_data_for_text:
        f.write(f"{name}\t{last_backup_created}\n")

print("Data successfully added to the text file.")

# Create a new Excel workbook and worksheet
output_wb = openpyxl.Workbook()
output_sheet = output_wb.active

# Write headers for the new Excel file
output_sheet.cell(row=1, column=1, value="Name")
output_sheet.cell(row=1, column=2, value="Last Backup Created")

# Write extracted data into the new worksheet
for row_index, (name, last_backup_created) in enumerate(extracted_data_for_text, start=2):
    output_sheet.cell(row=row_index, column=1, value=name)
    output_sheet.cell(row=row_index, column=2, value=last_backup_created)

# Save the new Excel workbook with extracted data
output_excel_path = "/home/serverguy/extracted_data.xlsx"
output_wb.save(output_excel_path)
print("Extracted data saved in Excel format.")

