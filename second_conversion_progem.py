#made by Shivam Sharma
import json
import openpyxl

json_file_path = "/home/serverguy/Downloads/file3.json"

# Read data from the JSON file
with open(json_file_path, "r") as json_file:
    data = json.load(json_file)["items"]

workbook = Workbook()
sheet = workbook.active

# Write headers (keys) as the first row
header_row = 1
headers = data[0].keys()
for col_index, key in enumerate(headers, start=1):
    sheet.cell(row=header_row, column=col_index, value=key)

# Write the JSON data to Excel rows
for row_index, item in enumerate(data, start=header_row + 1):
    for col_index, value in enumerate(item.values(), start=1):
        if isinstance(value, list):  # Handle list values
            value = ', '.join(map(str, value))  # Convert list items to strings
        elif isinstance(value, dict):  # Handle dictionary values
            value = json.dumps(value)  # Convert dictionary to JSON string
        sheet.cell(row=row_index, column=col_index, value=value)

output_excel_path = "/home/serverguy/today2_output.xlsx"
workbook.save(output_excel_path)
